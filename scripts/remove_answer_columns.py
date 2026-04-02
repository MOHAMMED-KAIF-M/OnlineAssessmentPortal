from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from questions.importers import find_header_row, normalize_key, normalize_text


ANSWER_HEADER_KEYS = {"answer", "correct answer", "expected answer", "expected_answer"}


def to_long_path(path: Path) -> str:
    resolved = path.resolve()
    return f"\\\\?\\{resolved}"


def answer_column_indexes(headers: list[object]) -> list[int]:
    indexes: list[int] = []
    for index, value in enumerate(headers, start=1):
        if normalize_key(value) in ANSWER_HEADER_KEYS:
            indexes.append(index)
    return indexes


def process_workbook(path: Path, dry_run: bool) -> tuple[bool, int]:
    workbook = load_workbook(to_long_path(path))
    changed = False
    removed_columns = 0
    try:
        for worksheet in workbook.worksheets:
            header_row = find_header_row(worksheet)
            if header_row is None:
                continue
            header_values = [
                normalize_text(value)
                for value in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
            ]
            indexes = answer_column_indexes(header_values)
            if not indexes:
                continue

            for column_index in reversed(indexes):
                worksheet.delete_cols(column_index, 1)
                removed_columns += 1
            changed = True

        if changed and not dry_run:
            workbook.save(to_long_path(path))
    finally:
        workbook.close()

    return changed, removed_columns


def matching_files(root: Path, contains: list[str], limit: int | None) -> list[Path]:
    files = sorted(root.rglob("*.xlsx"))
    if contains:
        lowered = [fragment.lower() for fragment in contains]
        files = [path for path in files if all(fragment in str(path).lower() for fragment in lowered)]
    if limit is not None:
        files = files[:limit]
    return files


def main() -> None:
    parser = argparse.ArgumentParser(description="Remove answer-bearing columns from workbook sheets.")
    parser.add_argument("--root", default="aq_files")
    parser.add_argument("--contains", nargs="*", default=[])
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    root = Path(args.root)
    files = matching_files(root, args.contains, args.limit)

    processed = 0
    changed_files = 0
    removed_columns = 0
    failures: list[tuple[Path, str]] = []

    for path in files:
        try:
            changed, removed = process_workbook(path, dry_run=args.dry_run)
            processed += 1
            changed_files += int(changed)
            removed_columns += removed
        except Exception as exc:  # noqa: BLE001
            failures.append((path, str(exc)))

    print(f"Processed: {processed}")
    print(f"Changed files: {changed_files}")
    print(f"Removed columns: {removed_columns}")
    print(f"Failures: {len(failures)}")
    if failures:
        print("Failure details:")
        for path, message in failures[:20]:
            print(f"- {path}: {message}")


if __name__ == "__main__":
    main()
