from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


HEADER_ROW_SCAN_LIMIT = 30


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def normalize_key(value: object) -> str:
    return normalize_text(value).lower()


def to_long_path(path: Path) -> str:
    resolved = path.resolve()
    return f"\\\\?\\{resolved}"


def header_has_difficulty(value: object) -> bool:
    return "difficulty" in normalize_key(value)


def split_csv_row(text: str) -> list[str]:
    return [part.strip() for part in next(csv.reader([text]))]


def join_csv_row(parts: list[str]) -> str:
    return ",".join(part.strip() for part in parts)


def header_score(values: list[object]) -> int:
    score = 0
    for value in values:
        key = normalize_key(value)
        if not key:
            continue
        if key in {"question", "question no.", "question no", "q.no", "question type", "type", "topic"}:
            score += 4
        elif "question" in key or "topic" in key:
            score += 2
        if header_has_difficulty(key):
            score += 3
    return score


def find_structured_header_row(ws) -> int | None:
    max_row = ws.max_row or 0
    best_row = None
    best_score = 0

    for row_no in range(1, min(max_row, HEADER_ROW_SCAN_LIMIT) + 1):
        values = [cell.value for cell in ws[row_no]]
        score = header_score(values)
        if score > best_score:
            best_score = score
            best_row = row_no

    return best_row if best_score >= 4 else None


def process_single_cell_sheet(ws) -> tuple[bool, int, int]:
    header = normalize_text(ws["A1"].value)
    if "," not in header:
        return False, 0, 0

    header_parts = split_csv_row(header)
    difficulty_indexes = [index for index, value in enumerate(header_parts) if header_has_difficulty(value)]
    if not difficulty_indexes:
        return False, 0, 0

    ws["A1"] = join_csv_row([value for index, value in enumerate(header_parts) if index not in difficulty_indexes])
    changed = True
    updated_rows = 0

    max_row = ws.max_row or 0
    for row_no in range(2, max_row + 1):
        raw_value = ws.cell(row=row_no, column=1).value
        text = normalize_text(raw_value)
        if not text:
            continue

        row_parts = split_csv_row(str(raw_value))
        if len(row_parts) <= max(difficulty_indexes):
            continue

        new_value = join_csv_row([value for index, value in enumerate(row_parts) if index not in difficulty_indexes])
        if new_value != str(raw_value):
            ws.cell(row=row_no, column=1).value = new_value
            updated_rows += 1

    return changed, len(difficulty_indexes), updated_rows


def process_structured_sheet(ws) -> tuple[bool, int, int]:
    header_row = find_structured_header_row(ws)
    if header_row is None:
        return False, 0, 0

    header_values = [cell.value for cell in ws[header_row]]
    difficulty_columns = [
        index
        for index, value in enumerate(header_values, start=1)
        if header_has_difficulty(value)
    ]
    if not difficulty_columns:
        return False, 0, 0

    for column_index in reversed(difficulty_columns):
        ws.delete_cols(column_index)

    return True, len(difficulty_columns), 0


def process_workbook(path: Path, dry_run: bool) -> tuple[bool, int, int, int]:
    workbook = load_workbook(to_long_path(path))
    workbook_changed = False
    removed_columns = 0
    updated_rows = 0
    changed_sheets = 0

    try:
        for ws in workbook.worksheets:
            single_changed, single_removed, single_rows = process_single_cell_sheet(ws)
            if single_changed:
                structured_changed, structured_removed = False, 0
            else:
                structured_changed, structured_removed, _ = process_structured_sheet(ws)
            sheet_changed = single_changed or structured_changed

            if sheet_changed:
                workbook_changed = True
                changed_sheets += 1
                removed_columns += single_removed + structured_removed
                updated_rows += single_rows

        if workbook_changed and not dry_run:
            workbook.save(to_long_path(path))
    finally:
        workbook.close()

    return workbook_changed, changed_sheets, removed_columns, updated_rows


def matching_files(root: Path) -> list[Path]:
    return sorted(root.rglob("*.xlsx"))


def main() -> None:
    parser = argparse.ArgumentParser(description="Remove difficulty columns from assessment workbook files.")
    parser.add_argument(
        "--roots",
        nargs="+",
        default=["Assessment Questions", "aq_files"],
        help="Root folders containing workbook files.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Analyze files without saving changes.",
    )
    args = parser.parse_args()

    processed_files = 0
    changed_files = 0
    changed_sheets = 0
    removed_columns = 0
    updated_rows = 0
    failures: list[tuple[Path, str]] = []

    for root_name in args.roots:
        root = Path(root_name)
        for path in matching_files(root):
            try:
                workbook_changed, workbook_sheets, workbook_columns, workbook_rows = process_workbook(
                    path,
                    dry_run=args.dry_run,
                )
                processed_files += 1
                changed_files += int(workbook_changed)
                changed_sheets += workbook_sheets
                removed_columns += workbook_columns
                updated_rows += workbook_rows
            except Exception as exc:  # noqa: BLE001
                failures.append((path, str(exc)))

    print(f"Processed files: {processed_files}")
    print(f"Changed files: {changed_files}")
    print(f"Changed sheets: {changed_sheets}")
    print(f"Removed difficulty columns: {removed_columns}")
    print(f"Updated CSV-style rows: {updated_rows}")
    print(f"Failures: {len(failures)}")

    if failures:
        print("Failure details:")
        for path, message in failures[:20]:
            print(f"- {path}: {message}")


if __name__ == "__main__":
    main()
