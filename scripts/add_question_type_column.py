from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from questions.question_types import normalize_question_type


QUESTION_VERBS = (
    "what",
    "which",
    "how",
    "why",
    "when",
    "where",
    "who",
    "create",
    "write",
    "build",
    "implement",
    "use",
    "calculate",
    "find",
    "sort",
    "fill",
    "clean",
    "compare",
    "demonstrate",
    "assign",
    "take",
    "convert",
    "swap",
    "display",
    "show",
    "determine",
    "identify",
    "analyze",
    "plot",
    "graph",
    "retrieve",
)

CODING_KEYWORDS = (
    "python",
    "code",
    "program",
    "function",
    "class",
    "lambda",
    "tf.keras",
    "keras",
    "tensorflow",
    "numpy",
    "pandas",
    "matplotlib",
    "seaborn",
    "algorithm",
    "script",
    "*args",
    "**kwargs",
)

SQL_KEYWORDS = (
    "sql",
    "query",
    "select ",
    " join ",
    "mysql",
    "database",
    "inner join",
    "left join",
    "right join",
    "group by",
    "having",
    "where clause",
)

PRACTICAL_KEYWORDS = (
    "vlookup",
    "index-match",
    "pivot",
    "dashboard",
    "excel",
    "tableau",
    "power bi",
    "chart",
    "line chart",
    "summary table",
    "git ",
    "command",
    "data cleaning",
    "clean the data",
    "break even",
    "contribution margin",
    "frequency table",
    "sort categories",
    "trend analysis",
)

SCENARIO_KEYWORDS = (
    "scenario",
    "customer",
    "company",
    "manager",
    "hospital",
    "retail",
    "business",
    "patient",
    "pharmacist",
    "receptionist",
    "ward manager",
    "sales manager",
    "department",
)

SHORT_ANSWER_KEYWORDS = (
    "one word",
    "two words",
    "one-word",
    "fill in the blank",
)

IGNORE_EXISTING_TYPES = {
    "data",
    "dataset",
    "header",
    "question",
    "questions",
    "type",
    "case1",
    "case2",
    "case3",
    "case4",
    "case5",
    "4.1",
    "4.2",
    "4.3",
    "4.4",
    "4.5",
}

OPTION_PATTERN = re.compile(r"\bA\)|\bB\)|\bC\)|\bD\)", re.IGNORECASE)
Q_NUMBER_PATTERN = re.compile(r"^\s*Q\d+[\.\,:]?\s*", re.IGNORECASE)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_key(value: object) -> str:
    return normalize_text(value).lower()


def to_long_path(path: Path) -> str:
    resolved = path.resolve()
    return f"\\\\?\\{resolved}"


def looks_like_question(text: str) -> bool:
    cleaned = normalize_text(text)
    if not cleaned:
        return False
    stripped = Q_NUMBER_PATTERN.sub("", cleaned).lower()
    if "?" in cleaned:
        return True
    if stripped.startswith(QUESTION_VERBS):
        return True
    return any(f" {verb} " in f" {stripped} " for verb in QUESTION_VERBS)


def standardize_existing_type(value: object) -> str | None:
    key = normalize_key(value)
    if not key or key in IGNORE_EXISTING_TYPES:
        return None
    normalized = normalize_question_type(value)
    return normalized or None


def infer_question_type(
    question_text: str,
    *,
    existing_type: object = None,
    options_present: bool = False,
    context: str = "",
) -> str | None:
    existing = standardize_existing_type(existing_type)
    if existing:
        return existing

    question_text = normalize_text(question_text)
    if not question_text and not options_present:
        return None

    combined = normalize_key(f"{question_text} {context}")

    if options_present or OPTION_PATTERN.search(question_text):
        return "MCQ"

    if any(token in combined for token in SHORT_ANSWER_KEYWORDS):
        return "Short Answer"

    if any(token in combined for token in SQL_KEYWORDS):
        return "SQL"

    if any(token in combined for token in CODING_KEYWORDS):
        return "Coding"

    if any(token in combined for token in SCENARIO_KEYWORDS) and (
        "?" in question_text or any(question_text.lower().startswith(verb) for verb in QUESTION_VERBS)
    ):
        return "Scenario-Based"

    if any(token in combined for token in PRACTICAL_KEYWORDS):
        return "Practical"

    lowered = question_text.lower()
    if lowered.startswith(("what is", "why", "how does", "define", "describe", "difference between", "explain")):
        return "Theory"

    if lowered.startswith(
        (
            "create",
            "write",
            "build",
            "implement",
            "use",
            "calculate",
            "find",
            "sort",
            "fill",
            "clean",
            "compare",
            "demonstrate",
            "assign",
            "take",
            "convert",
            "swap",
            "plot",
            "graph",
        )
    ):
        return "Practical"

    if looks_like_question(question_text):
        return "Theory"

    return None


def is_single_cell_layout(ws) -> bool:
    original_max_column = ws.max_column
    a1 = normalize_text(ws["A1"].value)
    b1 = normalize_text(ws.cell(row=1, column=2).value)
    if original_max_column == 1:
        return True
    return bool(a1 and b1 == "Question Type" and ("," in a1 or a1 == "Question"))


def extract_section_type(raw: str) -> str | None:
    first = raw.find(",")
    second = raw.find(",", first + 1)
    if first == -1 or second == -1:
        return None
    return raw[first + 1 : second]


def classify_single_cell_row(header: str, raw: object, context: str) -> str | None:
    text = normalize_text(raw)
    if not text:
        return None

    header_key = normalize_key(header)

    if header_key == "question":
        return infer_question_type(text, context=context)

    if header_key in {"q.no,question,type,difficulty", "q.no,question,type"}:
        if header_key == "q.no,question,type,difficulty":
            parts = text.rsplit(",", 2)
            if len(parts) == 3 and "," in parts[0]:
                _, question_text = parts[0].split(",", 1)
                return infer_question_type(question_text, existing_type=parts[1], context=context)
        else:
            parts = text.rsplit(",", 1)
            if len(parts) == 2 and "," in parts[0]:
                _, question_text = parts[0].split(",", 1)
                return infer_question_type(question_text, existing_type=parts[1], context=context)
        return infer_question_type(text, context=context)

    if header_key == "section,type,q.no,scenario,question,option a,option b,option c,option d,correct answer":
        return infer_question_type(
            question_text=text,
            existing_type=extract_section_type(text),
            options_present=True,
            context=context,
        )

    if header_key.startswith("type,"):
        parts = text.split(",", 3)
        if not parts or normalize_key(parts[0]) != "question":
            return None
        question_text = parts[3] if len(parts) == 4 else text
        question_text = re.sub(r"(,\s*)+$", "", question_text)
        return infer_question_type(question_text, context=context)

    if header_key.startswith("subtopic,scenario,type,"):
        parts = text.split(",", 7)
        if len(parts) < 8 or normalize_key(parts[2]) != "question":
            return None
        return infer_question_type(parts[7], context=f"{context} {parts[1]}")

    if header_key.startswith("scenario,"):
        if not re.match(r"^\s*q\d+\b", text, re.IGNORECASE):
            return None
        question_text = text.split(",", 1)[1] if "," in text else Q_NUMBER_PATTERN.sub("", text)
        return infer_question_type(question_text, context=f"{context} scenario")

    if header_key in {
        "month,sales ($)",
        "month,ad spend ($),sales ($)",
        "quarter,year 1,year 2,year 3",
        "dataset,values",
        "issue type,number of incidents",
        "reason,number of delays",
        "product code,annual sales value ($)",
    }:
        if not re.match(r"^\s*q\d+", text, re.IGNORECASE):
            return None
        question_text = Q_NUMBER_PATTERN.sub("", text)
        return infer_question_type(question_text, context=f"{context} data analysis")

    return infer_question_type(text, context=context)


def header_score(values: Iterable[object]) -> int:
    score = 0
    for value in values:
        key = normalize_key(value)
        if not key:
            continue
        if key == "question type":
            score += 6
        elif key in {"q.no", "question no.", "question no", "practical question"}:
            score += 5
        elif "sql question" in key:
            score += 5
        elif key == "question":
            score += 4
        elif key == "type":
            score += 3
        elif key.startswith("option ") or key in {"correct answer", "answer", "difficulty", "details / options"}:
            score += 2
        elif "question" in key:
            score += 2
    return score


def find_structured_header_row(ws) -> int | None:
    limit = min(ws.max_row, 30)
    best_row = None
    best_score = 0
    for row_no in range(1, limit + 1):
        values = next(ws.iter_rows(min_row=row_no, max_row=row_no, values_only=True))
        score = header_score(values)
        if score > best_score:
            best_score = score
            best_row = row_no
    return best_row if best_score >= 4 else None


def first_non_empty_question_text(row_values: tuple[object, ...], question_cols: list[int]) -> str:
    for col_idx in question_cols:
        if col_idx - 1 < len(row_values):
            value = normalize_text(row_values[col_idx - 1])
            if value:
                return value
    return ""


def process_single_cell_sheet(ws, context: str) -> tuple[bool, Counter]:
    changed = False
    counts: Counter[str] = Counter()
    header = normalize_text(ws["A1"].value)
    if ws.cell(row=1, column=2).value != "Question Type":
        ws.cell(row=1, column=2).value = "Question Type"
        changed = True

    for row_no in range(2, ws.max_row + 1):
        value = ws.cell(row=row_no, column=1).value
        question_type = classify_single_cell_row(header, value, context)
        current = normalize_text(ws.cell(row=row_no, column=2).value)
        target = question_type or ""
        if current != target:
            ws.cell(row=row_no, column=2).value = target or None
            changed = True
        if question_type:
            counts[question_type] += 1
    return changed, counts


def process_structured_sheet(ws, context: str) -> tuple[bool, Counter]:
    changed = False
    counts: Counter[str] = Counter()

    header_row = find_structured_header_row(ws)
    if header_row is None:
        return False, counts

    header_values = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    header_keys = [normalize_key(value) for value in header_values]

    if "question type" in header_keys:
        question_type_col = header_keys.index("question type") + 1
    else:
        question_type_col = ws.max_column + 1
        ws.cell(row=header_row, column=question_type_col).value = "Question Type"
        changed = True

    existing_type_col = header_keys.index("type") + 1 if "type" in header_keys else None
    question_cols = [
        idx
        for idx, key in enumerate(header_keys, start=1)
        if "question" in key and key not in {"question type", "question no.", "question no", "q.no"}
    ]
    options_present = any(
        key.startswith("option ") or key in {"correct answer", "answer", "details / options"} for key in header_keys
    )

    for row_no in range(header_row + 1, ws.max_row + 1):
        row_values = tuple(cell.value for cell in ws[row_no])
        question_text = first_non_empty_question_text(row_values, question_cols)
        existing_type = row_values[existing_type_col - 1] if existing_type_col and existing_type_col - 1 < len(row_values) else None

        if not question_text:
            continue

        if not looks_like_question(question_text) and not options_present:
            continue

        source_text = " ".join(normalize_text(value) for value in row_values if value not in (None, ""))
        question_type = infer_question_type(
            question_text=question_text,
            existing_type=existing_type,
            options_present=options_present,
            context=f"{context} {source_text}",
        )
        if not question_type:
            continue

        current = normalize_text(ws.cell(row=row_no, column=question_type_col).value)
        if current != question_type:
            ws.cell(row=row_no, column=question_type_col).value = question_type
            changed = True
        counts[question_type] += 1

    return changed, counts


def matching_files(root: Path, contains: list[str], limit: int | None) -> list[Path]:
    files = sorted(root.rglob("*.xlsx"))
    if contains:
        lowered = [fragment.lower() for fragment in contains]
        files = [path for path in files if all(fragment in str(path).lower() for fragment in lowered)]
    if limit is not None:
        files = files[:limit]
    return files


def process_workbook(path: Path, dry_run: bool) -> tuple[bool, Counter]:
    workbook = load_workbook(to_long_path(path))
    worksheet = workbook[workbook.sheetnames[0]]
    context = f"{path.parent.name} {path.stem}"

    if is_single_cell_layout(worksheet):
        changed, counts = process_single_cell_sheet(worksheet, context)
    else:
        changed, counts = process_structured_sheet(worksheet, context)

    if changed and not dry_run:
        workbook.save(to_long_path(path))
    workbook.close()
    return changed, counts


def main() -> None:
    parser = argparse.ArgumentParser(description="Add or fill a Question Type column across assessment workbooks.")
    parser.add_argument(
        "--root",
        default="Assessment Questions",
        help="Root folder containing workbook files.",
    )
    parser.add_argument(
        "--contains",
        nargs="*",
        default=[],
        help="Only process paths containing all provided fragments.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Process only the first N matching workbooks.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Analyze files without saving changes.",
    )
    args = parser.parse_args()

    root = Path(args.root)
    files = matching_files(root, args.contains, args.limit)

    changed_files = 0
    processed_files = 0
    type_counts: Counter[str] = Counter()
    failures: list[tuple[Path, str]] = []

    for path in files:
        try:
            changed, counts = process_workbook(path, dry_run=args.dry_run)
            processed_files += 1
            changed_files += int(changed)
            type_counts.update(counts)
        except Exception as exc:  # noqa: BLE001
            failures.append((path, str(exc)))

    print(f"Processed: {processed_files}")
    print(f"Changed: {changed_files}")
    print(f"Failures: {len(failures)}")
    for question_type, count in type_counts.most_common():
        print(f"{question_type}: {count}")

    if failures:
        print("Failure details:")
        for path, message in failures[:20]:
            print(f"- {path}: {message}")


if __name__ == "__main__":
    main()
