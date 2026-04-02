from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from questions.importers import (
    WorkbookMeta,
    infer_type_from_values,
    load_manifest,
    looks_like_question,
    normalize_key,
    normalize_text,
    split_csv_like_row,
    standardize_type,
)
from questions.question_types import normalize_question_type


TARGET_QUESTION_COUNT = 20
STANDARD_HEADERS = (
    'Question No.',
    'Topic',
    'Question Type',
    'Question',
    'Scenario/Context',
    'Skills Tested',
    'Option A',
    'Option B',
    'Option C',
    'Option D',
)
QUESTION_HEADERS = (
    'question',
    'write query',
    'practical question',
    'scenario-based sql question',
    'problem / question',
    'scenario question',
    'sql question',
    'query',
)
TOPIC_HEADERS = ('topic', 'subtopic', 'analytics_concept', 'concept')
CONTEXT_HEADERS = (
    'scenario/context',
    'scenario',
    'context',
    'analysis_required',
    'analysis required',
)
SKILL_HEADERS = (
    'skills tested',
    'key_learning',
    'key learning',
    'excel formula/setup',
    'excel setup',
    'excel setup/formula',
    'python code/skills',
)
OPTION_HEADERS = ('option a', 'option b', 'option c', 'option d')
MAX_HEADER_SCAN = 40
PLACEHOLDER_TYPES = {
    'technical': 'Theory',
    'conceptual': 'Theory',
    'application': 'Practical',
    'project': 'Practical',
    'design': 'Practical',
}


@dataclass
class ExtractedQuestion:
    question_text: str
    question_type: str = ''
    topic: str = ''
    context: str = ''
    skills_tested: str = ''
    option_a: str = ''
    option_b: str = ''
    option_c: str = ''
    option_d: str = ''
    correct_answer: str = ''


def to_long_path(path: Path) -> str:
    resolved = path.resolve()
    return f"\\\\?\\{resolved}"


def question_sheet(workbook):
    if 'questions' in workbook.sheetnames:
        return workbook['questions']
    return workbook[workbook.sheetnames[-1]]


def is_header_like(text: str) -> bool:
    key = normalize_key(text)
    return key in {
        'question',
        'question no',
        'question no.',
        'question type',
        'write query',
        'type',
    }


def first_question_header_index(header_keys: list[str]) -> int | None:
    for candidate in QUESTION_HEADERS:
        if candidate in header_keys:
            return header_keys.index(candidate)
    return None


def header_score(values: list[object]) -> int:
    keys = [normalize_key(value) for value in values]
    score = 0
    if first_question_header_index(keys) is not None:
        score += 6
    if 'question type' in keys or 'type' in keys:
        score += 4
    if any(key in {'question no', 'question no.', 'q.no', 'qno'} for key in keys):
        score += 3
    if any(key in TOPIC_HEADERS for key in keys):
        score += 2
    return score


def find_structured_header_row(worksheet) -> int | None:
    max_row = worksheet.max_row or 0
    best_row = None
    best_score = 0

    for row_number in range(1, min(max_row, MAX_HEADER_SCAN) + 1):
        values = [cell.value for cell in worksheet[row_number]]
        score = header_score(values)
        if score > best_score:
            best_score = score
            best_row = row_number

    return best_row if best_score >= 6 else None


def extract_from_structured_sheet(worksheet) -> list[ExtractedQuestion]:
    header_row = find_structured_header_row(worksheet)
    if header_row is None:
        return []

    header_values = [normalize_text(cell.value) for cell in worksheet[header_row]]
    header_keys = [normalize_key(value) for value in header_values]
    question_index = first_question_header_index(header_keys)
    if question_index is None:
        return []

    type_index = header_keys.index('question type') if 'question type' in header_keys else header_keys.index('type') if 'type' in header_keys else None
    topic_index = next((index for index, key in enumerate(header_keys) if key in TOPIC_HEADERS), None)
    context_indexes = [index for index, key in enumerate(header_keys) if key in CONTEXT_HEADERS]
    skill_indexes = [index for index, key in enumerate(header_keys) if key in SKILL_HEADERS]
    option_indexes = {
        label: header_keys.index(label) if label in header_keys else None
        for label in OPTION_HEADERS
    }
    rows: list[ExtractedQuestion] = []
    max_row = worksheet.max_row or 0
    for row_number in range(header_row + 1, max_row + 1):
        values = [cell.value for cell in worksheet[row_number]]
        question_text = normalize_text(values[question_index] if question_index < len(values) else '')
        if not question_text or is_header_like(question_text):
            continue

        topic = normalize_text(values[topic_index]) if topic_index is not None and topic_index < len(values) else ''
        explicit_type = values[type_index] if type_index is not None and type_index < len(values) else ''
        contexts = [
            normalize_text(values[index])
            for index in context_indexes
            if index < len(values) and normalize_text(values[index])
        ]
        skills = [
            normalize_text(values[index])
            for index in skill_indexes
            if index < len(values) and normalize_text(values[index])
        ]
        options = {
            key[-1].upper(): normalize_text(values[index]) if index is not None and index < len(values) else ''
            for key, index in option_indexes.items()
        }
        has_options = any(options.values())
        details = ' | '.join(contexts + skills)
        question_type = infer_type_from_values(
            question_text=question_text,
            explicit_type=explicit_type,
            header_question_label=header_values[question_index],
            details=details,
            has_options=has_options,
        )
        if not explicit_type and normalize_key(header_values[question_index]) in {'write query', 'sql question', 'query'}:
            question_type = 'SQL'
        rows.append(
            ExtractedQuestion(
                question_text=question_text,
                question_type=question_type,
                topic=topic,
                context=' | '.join(contexts),
                skills_tested=' | '.join(skills),
                option_a=options['A'],
                option_b=options['B'],
                option_c=options['C'],
                option_d=options['D'],
            )
        )
    return rows


def extract_from_two_column_sheet(worksheet) -> list[ExtractedQuestion]:
    if (worksheet.max_column or 0) < 2:
        return []

    second_header = normalize_key(worksheet.cell(row=1, column=2).value)
    first_cell = normalize_text(worksheet.cell(row=1, column=1).value)
    if second_header != 'question type' or not first_cell or is_header_like(first_cell):
        return []

    rows: list[ExtractedQuestion] = []
    max_row = worksheet.max_row or 0
    for row_number in range(1, max_row + 1):
        question_text = normalize_text(worksheet.cell(row=row_number, column=1).value)
        type_value = normalize_text(worksheet.cell(row=row_number, column=2).value)
        if not question_text:
            continue
        rows.append(
            ExtractedQuestion(
                question_text=question_text,
                question_type=(
                    infer_type_from_values(
                        question_text=question_text,
                        explicit_type=type_value,
                        header_question_label='question',
                        details='',
                        has_options=False,
                    )
                    if normalize_key(type_value) != 'question type'
                    else ''
                ),
            )
        )
    return rows


def header_indexes_from_csv(header_parts: list[str]) -> dict[str, object]:
    header_keys = [normalize_key(part) for part in header_parts]
    question_index = first_question_header_index(header_keys)
    topic_index = next((index for index, key in enumerate(header_keys) if key in TOPIC_HEADERS), None)
    type_index = header_keys.index('question type') if 'question type' in header_keys else header_keys.index('type') if 'type' in header_keys else None
    context_indexes = [index for index, key in enumerate(header_keys) if key in CONTEXT_HEADERS]
    skill_indexes = [index for index, key in enumerate(header_keys) if key in SKILL_HEADERS]
    option_indexes = {
        label: header_keys.index(label) if label in header_keys else None
        for label in OPTION_HEADERS
    }
    return {
        'question_index': question_index,
        'topic_index': topic_index,
        'type_index': type_index,
        'context_indexes': context_indexes,
        'skill_indexes': skill_indexes,
        'option_indexes': option_indexes,
    }


def extract_from_csv_single_column(worksheet) -> list[ExtractedQuestion]:
    if (worksheet.max_column or 0) != 1:
        return []

    header = normalize_text(worksheet.cell(row=1, column=1).value)
    if ',' not in header:
        return []

    header_parts = split_csv_like_row(header)
    indexes = header_indexes_from_csv(header_parts)
    question_index = indexes['question_index']
    if question_index is None:
        return []

    rows: list[ExtractedQuestion] = []
    max_row = worksheet.max_row or 0
    for row_number in range(2, max_row + 1):
        raw_value = normalize_text(worksheet.cell(row=row_number, column=1).value)
        if not raw_value:
            continue
        parts = split_csv_like_row(raw_value)
        if question_index >= len(parts):
            continue
        question_text = normalize_text(parts[question_index])
        if not question_text:
            continue

        topic = normalize_text(parts[indexes['topic_index']]) if indexes['topic_index'] is not None and indexes['topic_index'] < len(parts) else ''
        explicit_type = parts[indexes['type_index']] if indexes['type_index'] is not None and indexes['type_index'] < len(parts) else ''
        contexts = [
            normalize_text(parts[index])
            for index in indexes['context_indexes']
            if index < len(parts) and normalize_text(parts[index])
        ]
        skills = [
            normalize_text(parts[index])
            for index in indexes['skill_indexes']
            if index < len(parts) and normalize_text(parts[index])
        ]
        options = {
            key[-1].upper(): normalize_text(parts[index]) if index is not None and index < len(parts) else ''
            for key, index in indexes['option_indexes'].items()
        }
        has_options = any(options.values())
        details = ' | '.join(contexts + skills)
        question_type = infer_type_from_values(
            question_text=question_text,
            explicit_type=explicit_type,
            header_question_label=header_parts[question_index],
            details=details,
            has_options=has_options,
        )
        rows.append(
            ExtractedQuestion(
                question_text=question_text,
                question_type=question_type,
                topic=topic,
                context=' | '.join(contexts),
                skills_tested=' | '.join(skills),
                option_a=options['A'],
                option_b=options['B'],
                option_c=options['C'],
                option_d=options['D'],
            )
        )
    return rows


def extract_from_simple_single_column(worksheet) -> list[ExtractedQuestion]:
    if (worksheet.max_column or 0) != 1:
        return []

    rows: list[ExtractedQuestion] = []
    max_row = worksheet.max_row or 0
    for row_number in range(1, max_row + 1):
        question_text = normalize_text(worksheet.cell(row=row_number, column=1).value)
        if not question_text or is_header_like(question_text):
            continue
        if looks_like_question(question_text):
            rows.append(ExtractedQuestion(question_text=question_text, question_type='Theory'))
    return rows


def extract_existing_questions(workbook_path: Path) -> tuple[str, list[ExtractedQuestion]]:
    workbook = load_workbook(to_long_path(workbook_path), read_only=True, data_only=True)
    try:
        worksheet = question_sheet(workbook)
        candidates = {
            'structured': extract_from_structured_sheet(worksheet),
            'two_column': extract_from_two_column_sheet(worksheet),
            'csv_single_column': extract_from_csv_single_column(worksheet),
            'simple_single_column': extract_from_simple_single_column(worksheet),
        }
        best_mode = max(candidates, key=lambda key: len(candidates[key]))
        return best_mode, candidates[best_mode]
    finally:
        workbook.close()


def raw_question_sheet_count(workbook_path: Path) -> int:
    workbook = load_workbook(to_long_path(workbook_path), read_only=True, data_only=True)
    try:
        worksheet = question_sheet(workbook)
        non_empty_rows = 0
        max_row = worksheet.max_row or 0
        for row_number in range(2, max_row + 1):
            values = next(worksheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True))
            if any(value not in (None, '') for value in values):
                non_empty_rows += 1
        return non_empty_rows
    finally:
        workbook.close()


def preferred_type_from_rows(rows: list[ExtractedQuestion], meta: WorkbookMeta) -> str:
    type_counts = Counter()
    for row in rows:
        row_type = normalize_question_type(row.question_type) or PLACEHOLDER_TYPES.get(normalize_key(row.question_type), '')
        if row_type:
            type_counts[row_type] += 1
    if type_counts:
        return type_counts.most_common(1)[0][0]

    combined = normalize_key(f'{meta.workbook_title} {meta.module_path}')
    if any(token in combined for token in ('sql', 'database', 'query', 'mongodb', 'mongo')):
        return 'SQL'
    if any(token in combined for token in ('python', 'flask', 'pyspark', 'spark', 'tensorflow', 'keras', 'api')):
        return 'Coding'
    if any(token in combined for token in ('excel', 'tableau', 'power bi', 'aws', 'azure', 'git')):
        return 'Practical'
    return 'Theory'


def normalized_question_key(text: str) -> str:
    return normalize_key(text).strip(' ?.')


def clean_rows(rows: list[ExtractedQuestion], meta: WorkbookMeta, preferred_type: str) -> list[ExtractedQuestion]:
    cleaned: list[ExtractedQuestion] = []
    seen: set[str] = set()
    default_topic = meta.workbook_title
    for row in rows:
        question_text = normalize_text(row.question_text)
        if not question_text or is_header_like(question_text):
            continue
        question_key = normalized_question_key(question_text)
        if not question_key or question_key in seen:
            continue
        seen.add(question_key)
        question_type = (
            normalize_question_type(row.question_type)
            or PLACEHOLDER_TYPES.get(normalize_key(row.question_type), '')
            or preferred_type
        )
        cleaned.append(
            ExtractedQuestion(
                question_text=question_text,
                question_type=question_type,
                topic=normalize_text(row.topic) or default_topic,
                context=normalize_text(row.context),
                skills_tested=normalize_text(row.skills_tested),
                option_a=normalize_text(row.option_a),
                option_b=normalize_text(row.option_b),
                option_c=normalize_text(row.option_c),
                option_d=normalize_text(row.option_d),
            )
        )
    return cleaned


def default_topic(meta: WorkbookMeta) -> str:
    return meta.workbook_title


def mcq_row(question_text: str, topic: str, context: str, skill: str, correct_statement: str) -> ExtractedQuestion:
    return ExtractedQuestion(
        question_text=question_text,
        question_type='MCQ',
        topic=topic,
        context=context,
        skills_tested=skill,
        option_a=correct_statement,
        option_b='An unrelated manual process with no connection to the topic',
        option_c='A storage-only mechanism that does not solve the stated task',
        option_d='A deprecated approach that ignores the required data or workflow',
    )


def theory_templates() -> list[tuple[str, str]]:
    return [
        ('What is {topic}?', 'Foundations'),
        ('Why is {topic} important in {module_label}?', 'Importance'),
        ('Explain the main workflow of {topic}.', 'Process Understanding'),
        ('List the key components involved in {topic}.', 'Component Awareness'),
        ('How is {topic} applied in a practical business scenario?', 'Applied Understanding'),
        ('What problem does {topic} help solve?', 'Problem Framing'),
        ('Describe the core steps required to complete {topic}.', 'Execution Flow'),
        ('What inputs are typically required before starting {topic}?', 'Input Readiness'),
        ('What outputs or results are expected from {topic}?', 'Output Interpretation'),
        ('How would you validate whether {topic} was completed correctly?', 'Validation'),
        ('What are common mistakes to avoid while working on {topic}?', 'Error Awareness'),
        ('How does {topic} improve decision-making or outcomes?', 'Business Value'),
        ('What assumptions should be checked before using {topic}?', 'Assumption Checking'),
        ('How would you explain {topic} to a beginner?', 'Concept Simplification'),
        ('What metrics or indicators are useful when evaluating {topic}?', 'Measurement'),
        ('How does {topic} connect with earlier modules in the course?', 'Module Integration'),
        ('What are the prerequisites for learning {topic}?', 'Prerequisites'),
        ('Compare {topic} with a simpler alternative approach.', 'Comparison'),
        ('When should {topic} not be used?', 'Limitations'),
        ('Summarize the most important learning point from {topic}.', 'Retention'),
        ('What is the first step in implementing {topic}?', 'Initiation'),
        ('How can {topic} be reviewed after completion?', 'Review'),
        ('What risks should be monitored during {topic}?', 'Risk Awareness'),
        ('What makes {topic} effective in real projects?', 'Project Relevance'),
    ]


def practical_templates() -> list[tuple[str, str]]:
    return [
        ('Create a basic workflow to demonstrate {topic}.', 'Workflow Setup'),
        ('Perform a step-by-step practical exercise using {topic}.', 'Hands-On Execution'),
        ('Build a small example that applies {topic} to sample data.', 'Example Building'),
        ('Demonstrate how to validate the output produced by {topic}.', 'Output Validation'),
        ('Configure the minimum setup required to start {topic}.', 'Configuration'),
        ('Show how to prepare the input data or resources for {topic}.', 'Input Preparation'),
        ('Execute {topic} and document each major step clearly.', 'Procedure Documentation'),
        ('Apply {topic} to solve a simple business problem.', 'Business Application'),
        ('Create a reusable template for repeating {topic}.', 'Template Creation'),
        ('Demonstrate how to troubleshoot a failed {topic} workflow.', 'Troubleshooting'),
        ('Perform {topic} and explain the expected result at each stage.', 'Expected Outcomes'),
        ('Create a checklist to verify that {topic} was performed correctly.', 'Quality Checks'),
        ('Use {topic} to compare two different scenarios or inputs.', 'Scenario Comparison'),
        ('Demonstrate how to optimize the workflow used in {topic}.', 'Optimization'),
        ('Perform {topic} using a constrained or minimal dataset.', 'Lightweight Execution'),
        ('Create a summary view or report after completing {topic}.', 'Result Reporting'),
        ('Demonstrate how to reuse outputs from {topic} in the next task.', 'Workflow Linking'),
        ('Apply {topic} in a timed exercise and record the outcome.', 'Timed Practice'),
        ('Demonstrate how to review and correct mistakes made during {topic}.', 'Correction Loop'),
        ('Build a final practical example that combines the main ideas of {topic}.', 'Capstone Practice'),
        ('Perform an end-to-end run of {topic} from setup to result.', 'End-to-End Flow'),
        ('Create a practical checklist for repeating {topic} independently.', 'Independent Practice'),
        ('Demonstrate how to communicate the result of {topic} to a stakeholder.', 'Communication'),
        ('Apply {topic} on a new case and compare the result with the baseline case.', 'Transfer Learning'),
    ]


def coding_templates() -> list[tuple[str, str]]:
    return [
        ('Write a program to demonstrate {topic}.', 'Programming Fundamentals'),
        ('Implement a reusable function for {topic}.', 'Function Design'),
        ('Write code that validates the input required for {topic}.', 'Input Validation'),
        ('Build a script that performs {topic} end to end.', 'Script Workflow'),
        ('Implement error handling for a solution based on {topic}.', 'Error Handling'),
        ('Write code to test whether the output of {topic} is correct.', 'Testing'),
        ('Create a modular version of {topic} using clear functions or classes.', 'Modular Design'),
        ('Write a program that logs the important steps in {topic}.', 'Logging'),
        ('Build a parameterized solution for {topic}.', 'Parameterization'),
        ('Write a clean and readable implementation of {topic}.', 'Code Quality'),
        ('Implement a version of {topic} that works on sample input data.', 'Sample Execution'),
        ('Write a function that compares two outputs generated from {topic}.', 'Result Comparison'),
        ('Build a script that reports summary metrics for {topic}.', 'Metric Reporting'),
        ('Write code to automate a repetitive task related to {topic}.', 'Automation'),
        ('Implement unit-test-style checks for a {topic} solution.', 'Verification'),
        ('Write a solution for {topic} using efficient control flow.', 'Control Flow'),
        ('Build a debugging-friendly version of the code for {topic}.', 'Debugging'),
        ('Write code that handles edge cases in {topic}.', 'Edge Cases'),
        ('Implement a small project component centered on {topic}.', 'Project Integration'),
        ('Write a final polished solution that documents the steps for {topic}.', 'Documentation'),
        ('Implement a version of {topic} that reads configuration values externally.', 'Configuration Management'),
        ('Write code that transforms raw data before applying {topic}.', 'Data Preparation'),
        ('Build a simple interface or wrapper around {topic}.', 'Interface Design'),
        ('Implement performance checks for a program that uses {topic}.', 'Performance Awareness'),
    ]


def sql_templates() -> list[tuple[str, str]]:
    return [
        ('Write an SQL query to list all records needed for {topic}.', 'SELECT Basics'),
        ('Write an SQL query to filter records relevant to {topic}.', 'WHERE Filtering'),
        ('Write an SQL query that sorts the result for {topic}.', 'ORDER BY'),
        ('Write an SQL query that counts records related to {topic}.', 'COUNT Aggregation'),
        ('Write an SQL query that groups the result set for {topic}.', 'GROUP BY'),
        ('Write an SQL query that calculates an average or summary for {topic}.', 'Aggregate Functions'),
        ('Write an SQL query to find the highest-value result in {topic}.', 'MAX / TOP Analysis'),
        ('Write an SQL query to find the lowest-value result in {topic}.', 'MIN Analysis'),
        ('Write an SQL query that joins two tables for {topic}.', 'JOIN Logic'),
        ('Write an SQL query using aliases to improve readability for {topic}.', 'Aliases'),
        ('Write an SQL query that uses a subquery for {topic}.', 'Subqueries'),
        ('Write an SQL query that uses HAVING with grouped data for {topic}.', 'HAVING Clause'),
        ('Write an SQL query to find duplicate records related to {topic}.', 'Duplicate Detection'),
        ('Write an SQL query that updates rows associated with {topic}.', 'UPDATE'),
        ('Write an SQL query that deletes rows based on a condition in {topic}.', 'DELETE'),
        ('Write an SQL query to create a view for {topic}.', 'Views'),
        ('Write an SQL query that uses CASE logic for {topic}.', 'CASE Expressions'),
        ('Write an SQL query that ranks rows for {topic}.', 'Window Functions'),
        ('Write an SQL query that returns only distinct values for {topic}.', 'DISTINCT'),
        ('Write an SQL query to validate data quality for {topic}.', 'Data Quality Checks'),
        ('Write an SQL query that combines filters, grouping, and ordering for {topic}.', 'Combined Clauses'),
        ('Write an SQL query that compares current and previous values for {topic}.', 'Comparative Analysis'),
        ('Write an SQL query that identifies missing matches in a join for {topic}.', 'Outer Join Checks'),
        ('Write an SQL query that summarizes the final output needed for {topic}.', 'Final Reporting'),
    ]


def scenario_templates() -> list[tuple[str, str]]:
    return [
        ('A team needs to apply {topic} quickly for a new project. What should be done first?', 'Project Kickoff'),
        ('A stakeholder asks for a reliable output from {topic}. How should the workflow be planned?', 'Stakeholder Alignment'),
        ('A dataset is incomplete before using {topic}. What is the best next step?', 'Data Readiness'),
        ('A result from {topic} looks inconsistent. How should it be investigated?', 'Issue Investigation'),
        ('A business team wants faster delivery using {topic}. What improvement would you suggest?', 'Process Improvement'),
        ('A project owner wants to measure the impact of {topic}. Which checks should be included?', 'Impact Measurement'),
        ('A new analyst is assigned to {topic}. What guidance should they receive first?', 'Onboarding'),
        ('A workflow based on {topic} fails during execution. What should be reviewed immediately?', 'Failure Recovery'),
        ('A manager wants a concise update on {topic}. What should be communicated?', 'Reporting'),
        ('A team must choose between two approaches for {topic}. How should the decision be made?', 'Decision Framework'),
        ('A client requests a validated outcome for {topic}. What evidence should be prepared?', 'Client Assurance'),
        ('A process using {topic} must be repeated weekly. How should it be standardized?', 'Operational Consistency'),
        ('A reviewer questions the assumptions behind {topic}. How should you respond?', 'Assumption Defense'),
        ('A project must scale the use of {topic}. What should be addressed first?', 'Scalability'),
        ('A team needs to hand over work related to {topic}. What documentation is essential?', 'Knowledge Transfer'),
        ('A deadline is shortened for a task using {topic}. How should priorities be adjusted?', 'Prioritization'),
        ('A result from {topic} conflicts with domain expectations. What is the next best action?', 'Result Review'),
        ('A training session is planned on {topic}. What example should be used to teach it effectively?', 'Training Design'),
        ('A stakeholder asks how {topic} reduces business risk. What explanation would you provide?', 'Risk Reduction'),
        ('A project retrospective reviews {topic}. Which lesson should be captured?', 'Retrospective Learning'),
        ('A team wants to automate a repetitive part of {topic}. What should be automated first?', 'Automation Planning'),
        ('A new data source is added to a workflow that uses {topic}. What should be checked?', 'Integration Readiness'),
        ('A team must compare two outcomes produced by {topic}. What criteria should guide the comparison?', 'Evaluation Criteria'),
        ('A project lead asks for one improvement to strengthen {topic}. What would you recommend?', 'Continuous Improvement'),
    ]


def short_answer_templates() -> list[tuple[str, str]]:
    return [
        ('Define {topic} in one line.', 'Quick Definition'),
        ('Name one key benefit of {topic}.', 'Benefit Recall'),
        ('Name one risk associated with {topic}.', 'Risk Recall'),
        ('Name one prerequisite for {topic}.', 'Prerequisite Recall'),
        ('Name one tool or method used in {topic}.', 'Tool Awareness'),
        ('Name one output produced by {topic}.', 'Output Recall'),
        ('Name one validation step for {topic}.', 'Validation Recall'),
        ('Name one business use case for {topic}.', 'Use Case Recall'),
        ('Name one input required before starting {topic}.', 'Input Recall'),
        ('Name one common mistake in {topic}.', 'Mistake Recall'),
        ('Name one metric used to evaluate {topic}.', 'Metric Recall'),
        ('Name one reason to document {topic}.', 'Documentation Recall'),
        ('Name one dependency that can affect {topic}.', 'Dependency Recall'),
        ('Name one review step after completing {topic}.', 'Review Recall'),
        ('Name one scenario where {topic} is useful.', 'Scenario Recall'),
        ('Name one limitation of {topic}.', 'Limitation Recall'),
        ('Name one quality check for {topic}.', 'Quality Recall'),
        ('Name one stakeholder who benefits from {topic}.', 'Stakeholder Recall'),
        ('Name one way to improve {topic}.', 'Improvement Recall'),
        ('Name one learning outcome from {topic}.', 'Learning Recall'),
        ('Name one challenge in scaling {topic}.', 'Scale Recall'),
        ('Name one reason to automate {topic}.', 'Automation Recall'),
        ('Name one sign that {topic} worked correctly.', 'Success Recall'),
        ('Name one follow-up action after {topic}.', 'Follow-Up Recall'),
    ]


def generated_rows(meta: WorkbookMeta, preferred_type: str) -> list[ExtractedQuestion]:
    topic = default_topic(meta)
    module_label = meta.module_path.split('/')[-1] if meta.module_path else meta.course_code
    context = f"Course {meta.course_code} | {meta.module_path or meta.workbook_title}"

    if preferred_type == 'MCQ':
        prompts = [
            ('Which statement best describes {topic}?', 'Concept Recognition', '{topic} is used to solve the stated problem effectively'),
            ('Which option is the most appropriate use of {topic}?', 'Use Case Selection', '{topic} is applied when the workflow requires its core capability'),
            ('Which outcome is expected after applying {topic} correctly?', 'Outcome Prediction', '{topic} should produce a consistent and reviewable result'),
            ('Which step should happen first when working with {topic}?', 'Process Ordering', 'Prepare the required inputs before executing the workflow'),
            ('Which item is most important to verify after using {topic}?', 'Quality Validation', 'Validate the result against the expected business or technical requirement'),
            ('Which benefit is most strongly associated with {topic}?', 'Benefit Identification', '{topic} improves the quality or efficiency of the solution'),
            ('Which scenario is the best fit for {topic}?', 'Scenario Matching', '{topic} fits the scenario that directly needs its core method'),
            ('Which action helps reduce errors in {topic}?', 'Error Prevention', 'Review the inputs and assumptions before finalizing the result'),
            ('Which statement about {topic} is correct?', 'Concept Checking', '{topic} depends on clear inputs, a defined workflow, and result validation'),
            ('Which deliverable is commonly produced from {topic}?', 'Deliverable Awareness', 'A result, report, or output aligned with the task objective'),
            ('Which resource is most useful before starting {topic}?', 'Preparation', 'Relevant data, requirements, or configuration needed for execution'),
            ('Which practice best supports repeatability in {topic}?', 'Repeatability', 'Use a consistent method and document the steps performed'),
            ('Which review step is best after completing {topic}?', 'Post-Execution Review', 'Check whether the final output matches the expected objective'),
            ('Which metric would be reasonable to monitor for {topic}?', 'Measurement', 'A metric that reflects quality, accuracy, or completion'),
            ('Which change would improve a workflow based on {topic}?', 'Improvement Planning', 'Simplify steps while preserving the required output quality'),
            ('Which statement best explains the value of {topic}?', 'Value Awareness', '{topic} converts inputs into a usable and decision-ready result'),
            ('Which risk should be checked before using {topic}?', 'Risk Awareness', 'Invalid assumptions or incomplete inputs that weaken the result'),
            ('Which team behavior strengthens work on {topic}?', 'Team Practice', 'Clear communication, review, and documentation'),
            ('Which scenario shows successful use of {topic}?', 'Success Criteria', 'The process completes correctly and produces a useful output'),
            ('Which statement best summarizes {topic}?', 'Summary Understanding', '{topic} is a structured way to complete a defined technical or business task'),
            ('Which option best describes the first validation for {topic}?', 'Validation Readiness', 'Confirm that the source inputs are complete and relevant'),
            ('Which option best reflects a common challenge in {topic}?', 'Challenge Recognition', 'Handling incomplete data or unclear requirements'),
            ('Which option best supports scaling work related to {topic}?', 'Scalability Awareness', 'Create standardized steps and reusable checks'),
            ('Which option best explains why {topic} matters in projects?', 'Project Relevance', 'It helps produce reliable outcomes that teams can act on'),
        ]
        return [
            mcq_row(question_text.format(topic=topic), topic, context, skill, correct.format(topic=topic))
            for question_text, skill, correct in prompts
        ]

    if preferred_type == 'Practical':
        source = practical_templates()
        type_name = 'Practical'
    elif preferred_type == 'Coding':
        source = coding_templates()
        type_name = 'Coding'
    elif preferred_type == 'SQL':
        source = sql_templates()
        type_name = 'SQL'
    elif preferred_type == 'Scenario-Based':
        source = scenario_templates()
        type_name = 'Scenario-Based'
    elif preferred_type == 'Short Answer':
        source = short_answer_templates()
        type_name = 'Short Answer'
    else:
        source = theory_templates()
        type_name = 'Theory'

    return [
        ExtractedQuestion(
            question_text=prompt.format(topic=topic, module_label=module_label),
            question_type=type_name,
            topic=topic,
            context=context,
            skills_tested=skill,
        )
        for prompt, skill in source
    ]


def repaired_rows(meta: WorkbookMeta, existing_rows: list[ExtractedQuestion]) -> tuple[str, list[ExtractedQuestion]]:
    preferred_type = preferred_type_from_rows(existing_rows, meta)
    cleaned = clean_rows(existing_rows, meta, preferred_type)
    cleaned = cleaned[:TARGET_QUESTION_COUNT]

    generated = generated_rows(meta, preferred_type)
    seen = {normalized_question_key(row.question_text) for row in cleaned}
    for row in generated:
        if len(cleaned) >= TARGET_QUESTION_COUNT:
            break
        key = normalized_question_key(row.question_text)
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(row)

    while len(cleaned) < TARGET_QUESTION_COUNT:
        fallback_index = len(cleaned) + 1
        cleaned.append(
            ExtractedQuestion(
                question_text=f'Explain how {meta.workbook_title} should be applied in a real project scenario {fallback_index}.',
                question_type=preferred_type,
                topic=default_topic(meta),
                context=f'Course {meta.course_code} | {meta.module_path or meta.workbook_title}',
                skills_tested='General Understanding',
            )
        )

    return preferred_type, cleaned[:TARGET_QUESTION_COUNT]


def write_questions_sheet(path: Path, rows: list[ExtractedQuestion], dry_run: bool) -> None:
    workbook = load_workbook(to_long_path(path))
    try:
        existing_sheet = question_sheet(workbook)
        existing_index = workbook.index(existing_sheet)
        new_sheet = workbook.create_sheet(title='__tmp_questions__', index=existing_index)
        new_sheet.append(STANDARD_HEADERS)

        for row_number, row in enumerate(rows, start=1):
            new_sheet.append(
                [
                    row_number,
                    row.topic,
                    row.question_type,
                    row.question_text,
                    row.context,
                    row.skills_tested,
                    row.option_a,
                    row.option_b,
                    row.option_c,
                    row.option_d,
                ]
            )

        workbook.remove(existing_sheet)
        new_sheet.title = 'questions'
        if not dry_run:
            workbook.save(to_long_path(path))
    finally:
        workbook.close()


def needs_rebuild(mode: str, rows: list[ExtractedQuestion], workbook_path: Path) -> tuple[bool, list[str]]:
    reasons: list[str] = []
    if len(rows) != TARGET_QUESTION_COUNT:
        reasons.append(f'count_{len(rows)}')
    raw_count = raw_question_sheet_count(workbook_path)
    if raw_count != TARGET_QUESTION_COUNT:
        reasons.append(f'raw_count_{raw_count}')
    if mode != 'structured':
        reasons.append(f'mode_{mode}')
    if any(not normalize_text(row.question_type) for row in rows):
        reasons.append('blank_type')
    if any(normalize_text(row.question_type) and not normalize_question_type(row.question_type) for row in rows):
        reasons.append('unsupported_type')
    if any(
        normalize_text(row.question_type)
        and normalize_question_type(row.question_type)
        and normalize_text(row.question_type) != normalize_question_type(row.question_type)
        for row in rows
    ):
        reasons.append('nonstandard_type')
    if any(not normalize_text(row.question_text) for row in rows):
        reasons.append('blank_question')
    return bool(reasons), reasons


def manifest_targets(manifest_path: Path) -> list[tuple[WorkbookMeta, Path, Path]]:
    manifest = load_manifest(manifest_path)
    targets: list[tuple[WorkbookMeta, Path, Path]] = []
    with manifest_path.open(newline='', encoding='utf-8') as handle:
        for row in csv.DictReader(handle):
            extracted_path = Path(row['extracted_file'])
            meta = manifest[extracted_path.name]
            targets.append((meta, Path(row['original_path']), extracted_path))
    return targets


def main() -> None:
    parser = argparse.ArgumentParser(description='Repair question sheets and normalize them to 20 questions.')
    parser.add_argument('--manifest', default='aq_manifest.csv')
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--limit', type=int, default=None)
    args = parser.parse_args()

    manifest_path = Path(args.manifest)
    targets = manifest_targets(manifest_path)
    if args.limit is not None:
        targets = targets[: args.limit]

    processed = 0
    rebuilt = 0
    already_good = 0
    reason_counts: Counter[str] = Counter()
    failures: list[tuple[str, str]] = []

    for meta, original_path, extracted_path in targets:
        try:
            mode, rows = extract_existing_questions(extracted_path)
            rebuild_required, reasons = needs_rebuild(mode, rows, extracted_path)
            processed += 1
            if not rebuild_required:
                already_good += 1
                continue

            preferred_type, final_rows = repaired_rows(meta, rows)
            for target_path in (original_path, extracted_path):
                write_questions_sheet(target_path, final_rows, dry_run=args.dry_run)

            rebuilt += 1
            reason_counts.update(reasons)
            reason_counts[f'preferred_{preferred_type}'] += 1
        except Exception as exc:  # noqa: BLE001
            failures.append((extracted_path.name, str(exc)))

    print(f'Processed workbooks: {processed}')
    print(f'Already good: {already_good}')
    print(f'Rebuilt workbooks: {rebuilt}')
    print(f'Failures: {len(failures)}')
    if reason_counts:
        print('Rebuild reasons:')
        for reason, count in reason_counts.most_common():
            print(f'- {reason}: {count}')
    if failures:
        print('Failure details:')
        for name, message in failures[:20]:
            print(f'- {name}: {message}')


if __name__ == '__main__':
    main()
