from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from questions.importers import find_header_row, normalize_key, parse_packed_question_row

STANDARD_HEADERS = (
    'Question No.',
    'Topic',
    'Question Type',
    'Question',
    'Scenario/Context',
    'Skills Tested',
    'Option A',
    'Option B',
    'Option C',
    'Option D',
)


def to_long_path(path: Path) -> str:
    resolved = path.resolve()
    return f"\\\\?\\{resolved}"


def single_column_rows(worksheet, *, topic: str) -> list[list[object]] | None:
    if (worksheet.max_column or 0) != 1:
        return None

    header_value = worksheet.cell(row=1, column=1).value
    packed_header = parse_packed_question_row(header_value)
    if not packed_header or not packed_header['is_header']:
        return None

    rows: list[list[object]] = []
    max_row = worksheet.max_row or 0
    for row_number in range(2, max_row + 1):
        packed_row = parse_packed_question_row(worksheet.cell(row=row_number, column=1).value)
        if not packed_row or packed_row['is_header']:
            continue

        if normalize_key(packed_row['raw_type']) == 'dataset':
            if normalize_key(packed_row['question_number']) == 'source':
                question_text = 'Dataset columns: Source, Spend ($), Impressions, Clicks, Conversions, Revenue ($)'
            else:
                question_text = (
                    f"Dataset row - Source: {packed_row['question_number']} | "
                    f"Spend ($): {packed_row['question_text']} | "
                    f"Impressions: {packed_row['option_a']} | "
                    f"Clicks: {packed_row['option_b']} | "
                    f"Conversions: {packed_row['option_c']} | "
                    f"Revenue ($): {packed_row['option_d']}"
                )
            rows.append(
                [
                    len(rows) + 1,
                    topic,
                    'Practical',
                    question_text,
                    packed_row['context'],
                    '',
                    '',
                    '',
                    '',
                    '',
                ]
            )
            continue

        rows.append(
            [
                len(rows) + 1,
                topic,
                packed_row['question_type'] or 'Practical',
                packed_row['question_text'],
                packed_row['context'],
                '',
                packed_row['option_a'],
                packed_row['option_b'],
                packed_row['option_c'],
                packed_row['option_d'],
            ]
        )
    return rows


def process_worksheet(worksheet) -> tuple[bool, int, int]:
    header_row = find_header_row(worksheet)
    if header_row is None:
        return False, 0, 0

    header_values = [cell.value for cell in worksheet[header_row]]
    header_map = {normalize_key(value): index + 1 for index, value in enumerate(header_values)}
    question_col = header_map.get('question')
    type_col = header_map.get('question type')
    qno_col = header_map.get('question no.')
    context_col = header_map.get('scenario/context')
    option_cols = {
        'A': header_map.get('option a'),
        'B': header_map.get('option b'),
        'C': header_map.get('option c'),
        'D': header_map.get('option d'),
    }
    if question_col is None or type_col is None:
        return False, 0, 0

    changed = False
    updated_rows = 0
    rows_to_delete: list[int] = []

    max_row = worksheet.max_row or 0
    for row_number in range(header_row + 1, max_row + 1):
        packed_row = parse_packed_question_row(worksheet.cell(row=row_number, column=question_col).value)
        if not packed_row:
            continue
        if packed_row['is_header']:
            rows_to_delete.append(row_number)
            changed = True
            continue
        if not packed_row['question_type']:
            continue

        worksheet.cell(row=row_number, column=type_col).value = packed_row['question_type']
        worksheet.cell(row=row_number, column=question_col).value = packed_row['question_text']
        if context_col and packed_row['context'] and not worksheet.cell(row=row_number, column=context_col).value:
            worksheet.cell(row=row_number, column=context_col).value = packed_row['context']
        for label, column_index in option_cols.items():
            if column_index is None:
                continue
            packed_value = packed_row[f'option_{label.lower()}']
            if packed_value and not worksheet.cell(row=row_number, column=column_index).value:
                worksheet.cell(row=row_number, column=column_index).value = packed_value
        changed = True
        updated_rows += 1

    for row_number in reversed(rows_to_delete):
        worksheet.delete_rows(row_number, 1)

    if changed and qno_col:
        max_row = worksheet.max_row or 0
        for row_number in range(header_row + 1, max_row + 1):
            worksheet.cell(row=row_number, column=qno_col).value = row_number - header_row

    return changed, updated_rows, len(rows_to_delete)


def matching_files(root: Path, contains: list[str], limit: int | None) -> list[Path]:
    files = sorted(root.rglob('*.xlsx'))
    if contains:
        lowered = [fragment.lower() for fragment in contains]
        files = [path for path in files if all(fragment in str(path).lower() for fragment in lowered)]
    if limit is not None:
        files = files[:limit]
    return files


def process_workbook(path: Path, dry_run: bool) -> tuple[bool, int, int]:
    workbook = load_workbook(to_long_path(path))
    changed = False
    updated_rows = 0
    deleted_rows = 0
    try:
        for worksheet in workbook.worksheets:
            rebuilt_rows = single_column_rows(worksheet, topic=path.stem)
            if rebuilt_rows is not None:
                worksheet.delete_rows(1, worksheet.max_row or 0)
                worksheet.append(STANDARD_HEADERS)
                for row in rebuilt_rows:
                    worksheet.append(row)
                changed = True
                updated_rows += len(rebuilt_rows)
                deleted_rows += 1
                continue

            worksheet_changed, worksheet_updated, worksheet_deleted = process_worksheet(worksheet)
            changed = changed or worksheet_changed
            updated_rows += worksheet_updated
            deleted_rows += worksheet_deleted
        if changed and not dry_run:
            workbook.save(to_long_path(path))
    finally:
        workbook.close()
    return changed, updated_rows, deleted_rows


def main() -> None:
    parser = argparse.ArgumentParser(description='Unpack packed question-sheet rows into standard columns.')
    parser.add_argument('--root', default='aq_files')
    parser.add_argument('--contains', nargs='*', default=[])
    parser.add_argument('--limit', type=int, default=None)
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    root = Path(args.root)
    files = matching_files(root, args.contains, args.limit)

    processed = 0
    changed_files = 0
    updated_rows = 0
    deleted_rows = 0
    failures: list[tuple[Path, str]] = []

    for path in files:
        try:
            changed, workbook_updates, workbook_deletes = process_workbook(path, dry_run=args.dry_run)
            processed += 1
            changed_files += int(changed)
            updated_rows += workbook_updates
            deleted_rows += workbook_deletes
        except Exception as exc:  # noqa: BLE001
            failures.append((path, str(exc)))

    print(f'Processed: {processed}')
    print(f'Changed files: {changed_files}')
    print(f'Updated rows: {updated_rows}')
    print(f'Deleted rows: {deleted_rows}')
    print(f'Failures: {len(failures)}')
    if failures:
        print('Failure details:')
        for path, message in failures[:20]:
            print(f'- {path}: {message}')


if __name__ == '__main__':
    main()
