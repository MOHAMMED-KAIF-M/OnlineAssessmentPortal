import csv
import os
from collections import Counter
from itertools import groupby
from pathlib import Path

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import Http404
from django.shortcuts import redirect, render
from openpyxl import load_workbook
from questions.importers import normalize_key, parse_packed_question_row
from questions.models import QuestionBankEntry


COURSE_CODES = ("AIE", "CDA", "CDS", "CDE")
ANSWER_HEADER_KEYS = {"answer", "correct answer", "expected answer", "expected_answer"}


def _assessment_questions_root():
    return Path(settings.BASE_DIR) / 'Assessment Questions'


def _long_path(path):
    resolved = Path(path).resolve(strict=False)
    if os.name != 'nt':
        return str(resolved)
    return f"\\\\?\\{resolved}"


def _workbook_exists(path):
    candidate = Path(path)
    if candidate.exists() and candidate.is_file():
        return True
    try:
        with open(_long_path(candidate), 'rb'):
            return True
    except OSError:
        return False


def _build_folder_node(path, course_root):
    child_folders = []
    files = []

    for child in sorted(path.iterdir(), key=lambda item: (item.is_file(), item.name.lower())):
        if child.is_dir():
            child_folders.append(_build_folder_node(child, course_root))
        elif child.suffix.lower() == '.xlsx':
            files.append(
                {
                    'name': child.stem,
                    'filename': child.name,
                    'relative_path': child.relative_to(course_root).as_posix(),
                }
            )

    return {
        'name': path.name,
        'relative_path': path.relative_to(course_root).as_posix(),
        'folders': child_folders,
        'files': files,
        'child_folder_count': len(child_folders),
        'folder_count': sum(folder['folder_count'] + 1 for folder in child_folders),
        'file_count': len(files) + sum(folder['file_count'] for folder in child_folders),
    }


def _empty_course(course_code):
    return {
        'code': course_code,
        'folders': [],
        'module_count': 0,
        'folder_count': 0,
        'file_count': 0,
    }


def _build_course_data(course_code):
    assessment_root = _assessment_questions_root()
    course_path = assessment_root / course_code

    if not course_path.exists() or not course_path.is_dir():
        return _empty_course(course_code)

    course_root = _build_folder_node(course_path, assessment_root)
    return {
        'code': course_code,
        'folders': course_root['folders'],
        'module_count': course_root['child_folder_count'],
        'folder_count': course_root['folder_count'],
        'file_count': course_root['file_count'],
    }


def _course_catalog():
    courses = [_build_course_data(course_code) for course_code in COURSE_CODES]
    return {
        'courses': courses,
        'course_count': len(courses),
        'total_folder_count': sum(course['folder_count'] for course in courses),
        'total_file_count': sum(course['file_count'] for course in courses),
    }


def _resolve_relative_path(course_code, relative_path):
    if not relative_path:
        raise Http404("Workbook path is required.")

    relative_workbook_path = Path(str(relative_path).replace('\\', '/'))
    if relative_workbook_path.is_absolute():
        raise Http404("Invalid workbook path.")

    parts = [part for part in relative_workbook_path.parts if part not in ('', '.')]
    if not parts or '..' in parts:
        raise Http404("Invalid workbook path.")
    if parts[0].upper() != course_code:
        raise Http404("Workbook does not belong to this course.")

    workbook_path = _assessment_questions_root().joinpath(*parts)
    if workbook_path.suffix.lower() != '.xlsx' or not _workbook_exists(workbook_path):
        raise Http404("Workbook not found.")

    return workbook_path, '/'.join(parts)


def _serialize_option_set(entry):
    options = []
    for label, value in (
        ('A', entry.option_a),
        ('B', entry.option_b),
        ('C', entry.option_c),
        ('D', entry.option_d),
    ):
        if value:
            options.append({'label': label, 'value': value})
    return options


def _group_entries_by_sheet(entries):
    sheet_groups = []
    for sheet_name, sheet_entries in groupby(entries, key=lambda entry: entry.sheet_name or 'Sheet'):
        rows = []
        for entry in sheet_entries:
            rows.append(
                {
                    'row_number': entry.row_number,
                    'question_text': entry.question_text,
                    'question_type': entry.question_type,
                    'difficulty': entry.difficulty,
                    'context': entry.context,
                    'details': entry.details,
                    'options': _serialize_option_set(entry),
                }
            )
        sheet_groups.append(
            {
                'name': sheet_name,
                'entries': rows,
                'question_count': len(rows),
            }
        )
    return sheet_groups


def _display_cell(value):
    if value is None:
        return ''
    return str(value).strip()


def _strip_answer_columns(rows):
    if not rows:
        return rows

    header = rows[0]
    answer_indexes = {
        index
        for index, value in enumerate(header)
        if _display_cell(value).lower() in ANSWER_HEADER_KEYS
    }
    if not answer_indexes:
        return rows

    return [
        [cell for index, cell in enumerate(row) if index not in answer_indexes]
        for row in rows
    ]


def _normalize_packed_question_rows(rows):
    if not rows:
        return rows

    header = rows[0]
    header_map = {normalize_key(value): index for index, value in enumerate(header)}
    question_index = header_map.get('question')
    type_index = header_map.get('question type')
    context_index = header_map.get('scenario/context')
    option_indexes = {
        'A': header_map.get('option a'),
        'B': header_map.get('option b'),
        'C': header_map.get('option c'),
        'D': header_map.get('option d'),
    }
    if question_index is None or type_index is None:
        return rows

    normalized_rows = [header]
    for row in rows[1:]:
        padded_row = list(row) + [''] * (len(header) - len(row))
        packed_row = parse_packed_question_row(padded_row[question_index])
        if not packed_row:
            normalized_rows.append(padded_row)
            continue
        if packed_row['is_header']:
            continue
        if not packed_row['question_type']:
            normalized_rows.append(padded_row)
            continue

        padded_row[type_index] = packed_row['question_type']
        padded_row[question_index] = packed_row['question_text']
        if context_index is not None and packed_row['context'] and not padded_row[context_index]:
            padded_row[context_index] = packed_row['context']
        for label, index in option_indexes.items():
            if index is None:
                continue
            packed_value = packed_row[f'option_{label.lower()}']
            if packed_value and not padded_row[index]:
                padded_row[index] = packed_value
        normalized_rows.append(padded_row)

    return normalized_rows


def _expand_compound_sheet_row(cells):
    trimmed_cells = list(cells)
    while trimmed_cells and trimmed_cells[-1] == '':
        trimmed_cells.pop()

    if not trimmed_cells:
        return []

    non_empty_values = [cell for cell in trimmed_cells if cell]
    if len(non_empty_values) > 2 or ',' not in trimmed_cells[0]:
        return trimmed_cells

    try:
        expanded_first_cell = next(csv.reader([trimmed_cells[0]]))
    except csv.Error:
        return trimmed_cells

    if len(expanded_first_cell) < 3:
        return trimmed_cells

    expanded_row = [_display_cell(value) for value in expanded_first_cell]
    expanded_row.extend(trimmed_cells[1:])
    while expanded_row and expanded_row[-1] == '':
        expanded_row.pop()
    return expanded_row


def _sheet_selector_label(index):
    return f'Sheet {index} View'


def _sheet_matches(selected_sheet_name, index, worksheet_name):
    if not selected_sheet_name:
        return False

    normalized_selected = str(selected_sheet_name).strip().lower()
    aliases = {
        worksheet_name.strip().lower(),
        f'sheet{index}',
        f'sheet {index}',
    }
    return normalized_selected in aliases


def _sheet_entry_aliases(index, worksheet_name, total_sheets):
    aliases = {
        worksheet_name.strip().lower(),
        f'sheet{index}',
        f'sheet {index}',
    }
    if total_sheets == 1 and index == 1:
        aliases.add('questions')
    elif total_sheets == 2:
        if index == 1:
            aliases.add('dataset')
        elif index == 2:
            aliases.add('questions')
    return aliases


def _read_workbook_sheets(workbook_path, grouped_entries, selected_sheet_name):
    workbook = load_workbook(_long_path(workbook_path), data_only=True, read_only=True)
    try:
        sheet_views = []
        total_sheets = len(workbook.worksheets)
        normalized_grouped_entries = {
            key.strip().lower(): value for key, value in grouped_entries.items()
        }
        for index, worksheet in enumerate(workbook.worksheets, start=1):
            rows = []
            max_width = 0
            for value_row in worksheet.iter_rows(values_only=True):
                cells = _expand_compound_sheet_row([_display_cell(value) for value in value_row])
                if not cells:
                    continue
                max_width = max(max_width, len(cells))
                rows.append(cells)

            padded_rows = [row + [''] * (max_width - len(row)) for row in rows]
            padded_rows = _strip_answer_columns(padded_rows)
            padded_rows = _normalize_packed_question_rows(padded_rows)
            imported_sheet = {'entries': [], 'question_count': 0}
            for alias in _sheet_entry_aliases(index, worksheet.title, total_sheets):
                if alias in normalized_grouped_entries:
                    imported_sheet = normalized_grouped_entries[alias]
                    break
            sheet_views.append(
                {
                    'index': index,
                    'name': worksheet.title,
                    'selector_label': _sheet_selector_label(index),
                    'row_count': len(padded_rows),
                    'header': padded_rows[0] if padded_rows else [],
                    'body_rows': padded_rows[1:] if len(padded_rows) > 1 else [],
                    'entries': imported_sheet['entries'],
                    'question_count': imported_sheet['question_count'],
                }
            )

        if not sheet_views:
            return sheet_views, None

        active_sheet = next(
            (
                sheet
                for sheet in sheet_views
                if _sheet_matches(selected_sheet_name, sheet['index'], sheet['name'])
            ),
            sheet_views[0],
        )
        for sheet in sheet_views:
            sheet['is_active'] = sheet['name'] == active_sheet['name']
        return sheet_views, active_sheet
    finally:
        workbook.close()


def _workbook_context(course_code, relative_path, selected_sheet_name=None):
    workbook_path, normalized_relative_path = _resolve_relative_path(course_code, relative_path)
    source_path = f'Assessment Questions/{normalized_relative_path}'
    entries = list(
        QuestionBankEntry.objects.filter(source_path=source_path).order_by('sheet_name', 'row_number')
    )
    workbook_name = workbook_path.stem
    path_parts = normalized_relative_path.split('/')
    module_path = ' / '.join(path_parts[1:-1]) if len(path_parts) > 2 else ''
    type_counts = Counter(entry.question_type or 'Unspecified' for entry in entries)
    grouped_sheets = _group_entries_by_sheet(entries)
    grouped_sheet_lookup = {sheet['name']: sheet for sheet in grouped_sheets}
    sheet_views, active_sheet = _read_workbook_sheets(workbook_path, grouped_sheet_lookup, selected_sheet_name)

    return {
        'course': _build_course_data(course_code),
        'workbook': {
            'name': workbook_name,
            'filename': workbook_path.name,
            'relative_path': normalized_relative_path,
            'source_path': source_path,
            'module_path': module_path,
            'question_count': len(entries),
            'sheet_count': len(sheet_views),
            'imported_sheet_count': len(grouped_sheets),
            'sheets': grouped_sheets,
            'sheet_views': sheet_views,
            'active_sheet': active_sheet,
            'type_counts': [
                {'label': label, 'count': count}
                for label, count in sorted(type_counts.items(), key=lambda item: item[0].lower())
            ],
        },
    }


@login_required
def course_list(request):
    if not request.user.is_portal_admin:
        return redirect('student_dashboard')

    context = _course_catalog()
    context['admin_page'] = 'courses'

    return render(
        request,
        'courses/course_list.html',
        context,
    )


@login_required
def course_detail(request, course_code):
    if not request.user.is_portal_admin:
        return redirect('student_dashboard')

    normalized_course_code = course_code.upper()
    if normalized_course_code not in COURSE_CODES:
        raise Http404("Course not found.")

    return render(
        request,
        'courses/course_detail.html',
        {
            'course': _build_course_data(normalized_course_code),
            'admin_page': 'courses',
        },
    )


@login_required
def course_workbook_detail(request, course_code):
    if not request.user.is_portal_admin:
        return redirect('student_dashboard')

    normalized_course_code = course_code.upper()
    if normalized_course_code not in COURSE_CODES:
        raise Http404("Course not found.")

    context = _workbook_context(
        normalized_course_code,
        request.GET.get('path'),
        request.GET.get('sheet'),
    )
    context['admin_page'] = 'courses'

    return render(
        request,
        'courses/workbook_detail.html',
        context,
    )
