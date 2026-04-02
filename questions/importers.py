from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from itertools import combinations
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .question_types import normalize_question_type


QUESTION_VERBS = (
    'what',
    'which',
    'how',
    'why',
    'when',
    'where',
    'who',
    'create',
    'write',
    'build',
    'implement',
    'use',
    'calculate',
    'find',
    'sort',
    'fill',
    'clean',
    'compare',
    'demonstrate',
    'assign',
    'take',
    'convert',
    'swap',
    'display',
    'show',
    'determine',
    'identify',
    'analyze',
    'plot',
    'graph',
    'retrieve',
    'explain',
)

SQL_KEYWORDS = (
    'sql',
    'query',
    'select ',
    ' join ',
    'mysql',
    'database',
    'group by',
    'having',
)

CODING_KEYWORDS = (
    'python',
    'code',
    'program',
    'function',
    'class',
    'script',
    'tensorflow',
    'keras',
    'numpy',
    'pandas',
)

PRACTICAL_KEYWORDS = (
    'practical',
    'command',
    'operation',
    'project',
    'excel',
    'tableau',
    'power bi',
    'dashboard',
    'git ',
)

SHORT_ANSWER_KEYWORDS = (
    'one word',
    'two words',
    'one-word',
    'fill in the blank',
)

ANSWER_HEADER_KEYS = {
    'answer',
    'correct answer',
    'expected answer',
    'expected_answer',
}

PACKED_QUESTION_ROW_HEADERS = (
    'section',
    'type',
    'q.no',
    'scenario',
    'question',
    'option a',
    'option b',
    'option c',
    'option d',
    'correct answer',
)

@dataclass(slots=True)
class WorkbookMeta:
    source_workbook_id: int | None
    source_file: str
    source_path: str
    course_code: str
    module_path: str
    workbook_title: str


def normalize_text(value: object) -> str:
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip()


def normalize_key(value: object) -> str:
    return normalize_text(value).lower()


def is_answer_header(value: object) -> bool:
    return normalize_key(value) in ANSWER_HEADER_KEYS


def sanitize_raw_payload(raw_payload: dict[str, object] | None) -> dict[str, object]:
    sanitized = dict(raw_payload or {})
    for key in list(sanitized):
        if is_answer_header(key):
            sanitized[key] = ''
    return sanitized


def split_csv_like_row(text: str) -> list[str]:
    reader = csv.reader([text])
    return [part.strip() for part in next(reader)]


def join_csv_segments(parts: list[str]) -> str:
    return ', '.join(normalize_text(part) for part in parts if normalize_text(part))


def option_partition_cost(groups: list[str]) -> int:
    cost = 0
    for group in groups:
        text = normalize_text(group)
        if not text:
            return 10_000
        words = text.split()
        if text[0].islower():
            cost += 6
        if words and words[0].lower() in {'and', 'or', 'but', 'because', 'to', 'of', 'with', 'for', 'in', 'on', 'by'}:
            cost += 8
        if len(words) == 1:
            cost += 2
        if text.endswith('?') or text.endswith(':'):
            cost += 5
    return cost


def partition_option_segments(segments: list[str], option_count: int = 4) -> list[str] | None:
    if len(segments) < option_count:
        return None
    if len(segments) == option_count:
        return [normalize_text(segment) for segment in segments]

    best_groups: list[str] | None = None
    best_cost: int | None = None
    boundaries = range(1, len(segments))
    for cuts in combinations(boundaries, option_count - 1):
        indexes = (0, *cuts, len(segments))
        groups = [
            join_csv_segments(segments[indexes[index] : indexes[index + 1]])
            for index in range(option_count)
        ]
        cost = option_partition_cost(groups)
        if best_cost is None or cost < best_cost:
            best_cost = cost
            best_groups = groups
    return best_groups


def parse_packed_question_row(text: object) -> dict[str, object] | None:
    cleaned = normalize_text(text)
    if ',' not in cleaned:
        return None

    parts = split_csv_like_row(cleaned)
    normalized_parts = tuple(normalize_key(part) for part in parts)
    if normalized_parts == PACKED_QUESTION_ROW_HEADERS:
        return {'is_header': True}

    if len(parts) < 5:
        return None

    raw_type = normalize_text(parts[1]) if len(parts) > 1 else ''
    normalized_type = normalize_question_type(raw_type)
    if not normalized_type and normalize_key(raw_type) != 'dataset':
        return None

    question_text = normalize_text(parts[4]) if len(parts) > 4 else ''
    option_a = normalize_text(parts[5]) if len(parts) > 5 else ''
    option_b = normalize_text(parts[6]) if len(parts) > 6 else ''
    option_c = normalize_text(parts[7]) if len(parts) > 7 else ''
    option_d = normalize_text(parts[8]) if len(parts) > 8 else ''

    if normalized_type in {'MCQ', 'Scenario-Based'} and len(parts) > len(PACKED_QUESTION_ROW_HEADERS):
        answer = normalize_text(parts[-1]) if normalize_key(parts[-1]) in {'a', 'b', 'c', 'd'} else ''
        tail = parts[4:-1] if answer else parts[4:]
        best_candidate: tuple[str, list[str], int] | None = None
        for question_end in range(1, len(tail) - 3):
            candidate_question = join_csv_segments(tail[:question_end])
            if not candidate_question.endswith(('?', ':')):
                continue
            candidate_options = partition_option_segments(tail[question_end:], option_count=4)
            if not candidate_options:
                continue
            candidate_cost = option_partition_cost(candidate_options)
            if best_candidate is None or candidate_cost < best_candidate[2]:
                best_candidate = (candidate_question, candidate_options, candidate_cost)
        if best_candidate is not None:
            question_text = best_candidate[0]
            option_a, option_b, option_c, option_d = best_candidate[1]

    section = normalize_text(parts[0]) if len(parts) > 0 else ''
    scenario = normalize_text(parts[3]) if len(parts) > 3 else ''
    return {
        'is_header': False,
        'raw_type': raw_type,
        'question_type': normalized_type,
        'question_number': normalize_text(parts[2]) if len(parts) > 2 else '',
        'question_text': question_text,
        'context': ' | '.join(part for part in (section, scenario) if part),
        'option_a': option_a,
        'option_b': option_b,
        'option_c': option_c,
        'option_d': option_d,
    }


def is_qno_question_type_header(header_key: str) -> bool:
    return header_key in {'q.no,question,type,difficulty', 'q.no,question,type'}


def standardize_type(value: object) -> str:
    return normalize_question_type(value)


def looks_like_question(text: str) -> bool:
    cleaned = normalize_text(text)
    if not cleaned:
        return False
    lowered = cleaned.lower()
    if '?' in cleaned:
        return True
    if lowered.startswith(QUESTION_VERBS):
        return True
    return any(f' {verb} ' in f' {lowered} ' for verb in QUESTION_VERBS)


def extract_options(text: str) -> dict[str, str]:
    cleaned = normalize_text(text)
    matches = list(re.finditer(r'([A-D])\)', cleaned, re.IGNORECASE))
    if len(matches) < 2:
        return {}

    options: dict[str, str] = {}
    for index, match in enumerate(matches):
        option_key = match.group(1).upper()
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(cleaned)
        options[option_key] = cleaned[start:end].strip(' ,;')
    return options


def infer_type_from_values(
    *,
    question_text: str,
    explicit_type: object = None,
    header_question_label: str = '',
    details: str = '',
    has_options: bool = False,
) -> str:
    explicit = standardize_type(explicit_type)
    if explicit:
        return explicit

    label = normalize_key(header_question_label)
    combined = normalize_key(f'{question_text} {details} {label}')

    if has_options:
        return 'MCQ'
    if any(token in combined for token in SHORT_ANSWER_KEYWORDS):
        return 'Short Answer'
    if any(token in combined for token in SQL_KEYWORDS):
        return 'SQL'
    if any(token in combined for token in CODING_KEYWORDS):
        return 'Coding'
    if any(token in combined for token in PRACTICAL_KEYWORDS):
        return 'Practical'
    if 'scenario' in combined:
        return 'Scenario-Based'
    if looks_like_question(question_text):
        return 'Theory'
    return 'Practical'


def load_manifest(manifest_path: Path) -> dict[str, WorkbookMeta]:
    manifest: dict[str, WorkbookMeta] = {}
    with manifest_path.open(newline='', encoding='utf-8') as handle:
        for row in csv.DictReader(handle):
            extracted_file = Path(row['extracted_file']).name
            source_path = row['original_path']
            path_parts = Path(source_path).parts
            course_code = path_parts[1] if len(path_parts) > 1 else ''
            module_path = '/'.join(path_parts[2:-1]) if len(path_parts) > 3 else ''
            workbook_title = Path(source_path).stem or Path(extracted_file).stem
            manifest[extracted_file] = WorkbookMeta(
                source_workbook_id=int(row['index']) if row.get('index') else None,
                source_file=extracted_file,
                source_path=source_path,
                course_code=course_code,
                module_path=module_path,
                workbook_title=workbook_title,
            )
    return manifest


def first_question_column(header_keys: list[str]) -> int | None:
    preferred = (
        'question',
        'practical question',
        'scenario-based sql question',
        'problem / question',
        'scenario question',
        'sql question',
    )
    for name in preferred:
        if name in header_keys:
            return header_keys.index(name)

    for index, key in enumerate(header_keys):
        if 'question' in key and key not in {'question no.', 'question no', 'q.no', 'question type'}:
            return index
    return None


def find_header_row(worksheet) -> int | None:
    limit = min(worksheet.max_row, 20)
    for row_number in range(1, limit + 1):
        values = [normalize_key(value) for value in next(worksheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True))]
        if first_question_column(values) is not None:
            return row_number
    return None


def question_rows_from_single_column(meta: WorkbookMeta, worksheet) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    header = normalize_text(worksheet.cell(row=1, column=1).value)
    header_key = normalize_key(header)

    if is_qno_question_type_header(header_key):
        start_row = 2
    else:
        start_row = 1

    for row_number in range(start_row, worksheet.max_row + 1):
        value = normalize_text(worksheet.cell(row=row_number, column=1).value)
        if not value:
            continue

        if is_qno_question_type_header(header_key):
            parts = split_csv_like_row(value)
            if header_key == 'q.no,question,type,difficulty':
                if len(parts) < 4:
                    continue
                question_text = ','.join(parts[1:-2]).strip()
                explicit_type = parts[-2]
                difficulty = parts[-1]
            else:
                if len(parts) < 3:
                    continue
                question_text = ','.join(parts[1:-1]).strip()
                explicit_type = parts[-1]
                difficulty = ''
            if not question_text:
                continue
            question_type = infer_type_from_values(
                question_text=question_text,
                explicit_type=explicit_type,
                header_question_label='question',
                details='',
                has_options=False,
            )
            rows.append(
                build_record(
                    meta,
                    worksheet.title,
                    row_number,
                    question_text=question_text,
                    question_type=question_type,
                    difficulty=difficulty,
                    raw_payload={'raw': value},
                )
            )
            continue

        question_type = ''
        if header_key == 'scenario question':
            question_type = 'Scenario-Based'
        elif not looks_like_question(value):
            continue

        rows.append(
            build_record(
                meta,
                worksheet.title,
                row_number,
                question_text=value,
                question_type=question_type or 'Theory',
                raw_payload={'raw': value},
            )
        )
    return rows


def question_rows_from_structured_sheet(meta: WorkbookMeta, worksheet) -> list[dict[str, object]]:
    header_row = find_header_row(worksheet)
    if header_row is None:
        return []

    header_values = [normalize_text(value) for value in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    header_keys = [normalize_key(value) for value in header_values]
    question_col = first_question_column(header_keys)
    if question_col is None:
        return []

    type_col = header_keys.index('question type') if 'question type' in header_keys else header_keys.index('type') if 'type' in header_keys else None
    topic_col = header_keys.index('topic') if 'topic' in header_keys else None
    difficulty_col = next((i for i, key in enumerate(header_keys) if 'difficulty' in key), None)
    option_map = {
        'A': header_keys.index('option a') if 'option a' in header_keys else None,
        'B': header_keys.index('option b') if 'option b' in header_keys else None,
        'C': header_keys.index('option c') if 'option c' in header_keys else None,
        'D': header_keys.index('option d') if 'option d' in header_keys else None,
    }

    rows: list[dict[str, object]] = []
    question_label = header_keys[question_col]

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        values = [cell.value for cell in worksheet[row_number]]
        question_text = normalize_text(values[question_col] if question_col < len(values) else '')
        if not question_text:
            continue

        topic = normalize_text(values[topic_col]) if topic_col is not None and topic_col < len(values) else ''
        explicit_type = values[type_col] if type_col is not None and type_col < len(values) else ''
        difficulty = normalize_text(values[difficulty_col]) if difficulty_col is not None and difficulty_col < len(values) else ''
        contexts: list[str] = []
        option_values = {
            key: normalize_text(values[index]) if index is not None and index < len(values) else ''
            for key, index in option_map.items()
        }
        has_options = any(option_values.values())

        packed_row = parse_packed_question_row(question_text) if not has_options else None
        if packed_row and packed_row['is_header']:
            continue
        if packed_row and normalize_key(packed_row['raw_type']) == 'dataset':
            continue
        if packed_row:
            question_text = packed_row['question_text']
            explicit_type = packed_row['question_type'] or explicit_type
            packed_context = packed_row['context']
            if packed_context:
                contexts = [packed_context, *contexts] if packed_context not in contexts else contexts
            for label in ('A', 'B', 'C', 'D'):
                packed_value = packed_row[f'option_{label.lower()}']
                if packed_value and not option_values[label]:
                    option_values[label] = packed_value
            has_options = any(option_values.values())
        if not question_text:
            continue

        details_parts: list[str] = []
        for index, key in enumerate(header_keys):
            if index in {
                question_col,
                type_col,
                topic_col,
                difficulty_col,
                *(idx for idx in option_map.values() if idx is not None),
            }:
                continue
            if is_answer_header(key):
                continue
            value = normalize_text(values[index]) if index < len(values) else ''
            if value:
                details_parts.append(f'{header_values[index]}: {value}')

        details = ' | '.join(details_parts)

        if not has_options:
            embedded_options = extract_options(details)
            if embedded_options:
                option_values.update(embedded_options)
                has_options = True

        question_type = infer_type_from_values(
            question_text=question_text,
            explicit_type=explicit_type,
            header_question_label=question_label,
            details=details,
            has_options=has_options,
        )

        rows.append(
            build_record(
                meta,
                worksheet.title,
                row_number,
                topic=topic,
                question_text=question_text,
                question_type=question_type,
                difficulty=difficulty,
                details=details,
                option_a=option_values['A'],
                option_b=option_values['B'],
                option_c=option_values['C'],
                option_d=option_values['D'],
                correct_answer='',
                raw_payload=sanitize_raw_payload(
                    dict(zip(header_values, (normalize_text(value) for value in values)))
                ),
            )
        )
    return rows


def build_record(
    meta: WorkbookMeta,
    sheet_name: str,
    row_number: int,
    *,
    question_text: str,
    question_type: str = '',
    difficulty: str = '',
    topic: str = '',
    context: str = '',
    details: str = '',
    option_a: str = '',
    option_b: str = '',
    option_c: str = '',
    option_d: str = '',
    correct_answer: str = '',
    raw_payload: dict[str, object] | None = None,
) -> dict[str, object]:
    return {
        'source_workbook_id': meta.source_workbook_id,
        'source_file': meta.source_file,
        'source_path': meta.source_path,
        'course_code': meta.course_code,
        'module_path': meta.module_path,
        'workbook_title': meta.workbook_title,
        'sheet_name': sheet_name,
        'row_number': row_number,
        'topic': topic,
        'question_type': question_type,
        'difficulty': difficulty,
        'question_text': question_text,
        'context': context,
        'details': details,
        'option_a': option_a,
        'option_b': option_b,
        'option_c': option_c,
        'option_d': option_d,
        'correct_answer': '',
        'raw_payload': json.dumps(sanitize_raw_payload(raw_payload), ensure_ascii=False),
    }


def parse_workbook(path: Path, meta: WorkbookMeta | None = None) -> list[dict[str, object]]:
    workbook_meta = meta or WorkbookMeta(
        source_workbook_id=None,
        source_file=path.name,
        source_path=str(path),
        course_code='',
        module_path='',
        workbook_title=path.stem,
    )

    workbook = load_workbook(path, data_only=True)
    try:
        rows: list[dict[str, object]] = []
        for worksheet in workbook.worksheets:
            if worksheet.max_column == 1:
                rows.extend(question_rows_from_single_column(workbook_meta, worksheet))
            else:
                rows.extend(question_rows_from_structured_sheet(workbook_meta, worksheet))
        return rows
    finally:
        workbook.close()


def workbook_matches(path: Path, contains: Iterable[str]) -> bool:
    lowered = str(path).lower()
    return all(fragment.lower() in lowered for fragment in contains)
