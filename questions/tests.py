import os
import json
from pathlib import Path
from tempfile import mkstemp

from django.test import SimpleTestCase
from openpyxl import Workbook

from .importers import (
    WorkbookMeta,
    extract_options,
    infer_type_from_values,
    parse_packed_question_row,
    parse_workbook,
    question_rows_from_single_column,
    split_csv_like_row,
    standardize_type,
)


class ImporterHelperTests(SimpleTestCase):
    def test_extract_options_from_mcq_text(self):
        options = extract_options('A) Apple B) Banana C) Cherry D) Date')

        self.assertEqual(options['A'], 'Apple')
        self.assertEqual(options['B'], 'Banana')
        self.assertEqual(options['C'], 'Cherry')
        self.assertEqual(options['D'], 'Date')

    def test_split_csv_like_row_rebuilds_middle_question_text(self):
        row = split_csv_like_row('Q1,What is SQL, and why is it used?,Theory,Easy')

        self.assertEqual(row[0], 'Q1')
        self.assertEqual(row[-2], 'Theory')
        self.assertEqual(row[-1], 'Easy')
        self.assertEqual(', '.join(row[1:-2]), 'What is SQL, and why is it used?')

    def test_infer_type_prefers_explicit_type(self):
        question_type = infer_type_from_values(
            question_text='Write a query to list all rows',
            explicit_type='SQL',
            header_question_label='question',
            details='',
            has_options=False,
        )

        self.assertEqual(question_type, 'SQL')

    def test_standardize_type_maps_project_to_practical(self):
        self.assertEqual(standardize_type('Project'), 'Practical')

    def test_infer_type_ignores_nonstandard_explicit_type_when_text_is_clear(self):
        question_type = infer_type_from_values(
            question_text='Write a SELECT query to list all rows',
            explicit_type='Unsupported',
            header_question_label='question',
            details='',
            has_options=False,
        )

        self.assertEqual(question_type, 'SQL')

    def test_single_column_import_supports_qno_question_type_without_difficulty(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Q.No,Question,Type'
        worksheet['A2'] = 'Q1,What is Python?,Theory'
        worksheet['A3'] = 'Q2,Write a SELECT query,SQL'
        meta = WorkbookMeta(
            source_workbook_id=1,
            source_file='0001.xlsx',
            source_path='Assessment Questions/AIE/Sample/Sample.xlsx',
            course_code='AIE',
            module_path='Sample',
            workbook_title='Sample',
        )

        rows = question_rows_from_single_column(meta, worksheet)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]['question_type'], 'Theory')
        self.assertEqual(rows[0]['difficulty'], '')
        self.assertEqual(rows[1]['question_type'], 'SQL')

    def test_single_column_import_normalizes_project_type(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Q.No,Question,Type'
        worksheet['A2'] = 'Q1,Complete an end-to-end ML project,Project'
        meta = WorkbookMeta(
            source_workbook_id=1,
            source_file='0001.xlsx',
            source_path='Assessment Questions/AIE/Sample/Sample.xlsx',
            course_code='AIE',
            module_path='Sample',
            workbook_title='Sample',
        )

        rows = question_rows_from_single_column(meta, worksheet)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['question_type'], 'Practical')

    def test_parse_workbook_omits_correct_answer_values(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'questions'
        worksheet.append(['Question', 'Option A', 'Option B', 'Option C', 'Option D', 'Correct Answer', 'Question Type'])
        worksheet.append(['What is AI bias?', 'Fast compute', 'Unfair outcomes', 'Data storage', 'Accuracy', 'B', 'MCQ'])

        fd, tmp_path = mkstemp(suffix='.xlsx')
        os.close(fd)
        path = Path(tmp_path)
        try:
            workbook.save(path)
            rows = parse_workbook(path)
        finally:
            if path.exists():
                path.unlink()

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['correct_answer'], '')
        self.assertEqual(json.loads(rows[0]['raw_payload'])['Correct Answer'], '')

    def test_parse_packed_question_row_recognizes_embedded_mcq(self):
        packed = parse_packed_question_row(
            'SECTION A: Medium Level MCQs,MCQ,1,,What is Business Analytics?,Option A,Option B,Option C,Option D,A'
        )

        self.assertIsNotNone(packed)
        self.assertEqual(packed['question_type'], 'MCQ')
        self.assertEqual(packed['question_text'], 'What is Business Analytics?')
        self.assertEqual(packed['option_b'], 'Option B')

    def test_parse_packed_question_row_preserves_commas_inside_option_text(self):
        packed = parse_packed_question_row(
            'SECTION A: Medium Level MCQs,MCQ,4,,Supply chain analytics helps businesses:,Increase delivery times,Optimize logistics, reduce costs, and improve efficiency,Hold more inventory,Eliminate suppliers,B'
        )

        self.assertIsNotNone(packed)
        self.assertEqual(packed['option_a'], 'Increase delivery times')
        self.assertEqual(
            packed['option_b'],
            'Optimize logistics, reduce costs, and improve efficiency',
        )
        self.assertEqual(packed['option_c'], 'Hold more inventory')
        self.assertEqual(packed['option_d'], 'Eliminate suppliers')

    def test_parse_workbook_unpacks_packed_structured_mcq_rows(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'questions'
        worksheet.append(['Question No.', 'Topic', 'Question Type', 'Question', 'Scenario/Context', 'Skills Tested', 'Option A', 'Option B', 'Option C', 'Option D'])
        worksheet.append([1, 'Business Analytics Overview', 'MCQ', 'Section,Type,Q.No,Scenario,Question,Option A,Option B,Option C,Option D,Correct Answer', '', '', '', '', '', ''])
        worksheet.append([2, 'Business Analytics Overview', 'MCQ', 'SECTION A: Medium Level MCQs,MCQ,1,,What is Business Analytics?,Raw data only,Using data for decision-making,Presentation design,Database storage,B', '', '', '', '', '', ''])

        fd, tmp_path = mkstemp(suffix='.xlsx')
        os.close(fd)
        path = Path(tmp_path)
        try:
            workbook.save(path)
            rows = parse_workbook(path)
        finally:
            if path.exists():
                path.unlink()

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['question_type'], 'MCQ')
        self.assertEqual(rows[0]['question_text'], 'What is Business Analytics?')
        self.assertEqual(rows[0]['option_b'], 'Using data for decision-making')

    def test_parse_workbook_reads_question_rows_from_later_sheet(self):
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = 'Sheet1'
        sheet1['A1'] = 'Analytics_ID'
        sheet1['B1'] = 'Question Type'
        sheet1['A2'] = 'BA-001'
        sheet1['B2'] = 'Coding'

        sheet2 = workbook.create_sheet('Sheet2')
        sheet2['A1'] = 'Q.No,Question,Type'
        sheet2['A2'] = 'Q1,What is Big Data Analytics?,Theory'

        meta = WorkbookMeta(
            source_workbook_id=344,
            source_file='0344.xlsx',
            source_path='Assessment Questions/CDA/BDF-117/BIG DATA INTRODUCTION/Big Data Analytics Introduction/Big Data Analytics Introduction.xlsx',
            course_code='CDA',
            module_path='BDF-117/BIG DATA INTRODUCTION/Big Data Analytics Introduction',
            workbook_title='Big Data Analytics Introduction',
        )

        fd, tmp_path = mkstemp(suffix='.xlsx')
        os.close(fd)
        path = Path(tmp_path)
        try:
            workbook.save(path)
            rows = parse_workbook(path, meta)
        finally:
            if path.exists():
                path.unlink()

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['sheet_name'], 'Sheet2')
        self.assertEqual(rows[0]['question_text'], 'What is Big Data Analytics?')
